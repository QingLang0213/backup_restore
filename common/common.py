#coding=utf-8
import random,time
import base64
import socket
import struct
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import  Font, Alignment
from getconfigs import GetConfigs



class Common(object):
    def __init__(self,ssh,r,cur,mysql_table_name,flag):
        self.ssh=ssh
        self.r=r
        self.cur=cur
        self.mysql_table_name=mysql_table_name
        self.flag=flag
        self.config = GetConfigs()
        
        #define array to store values
        self.mysql_keys=[]
        self.redis_keys_list=[]
        self.redis_values_list=[]
        self.mysql_values_list=[]
        self.mysql_keys_list=[]
        
    def set_restore_name(self):
        self.restore_info=self.config.getstr(self.mysql_table_name,'restore_info','restore_table')
        table_date=self.restore_info.split()[0]
        table_set=self.restore_info.split()[-1]
        ymd=time.strftime("%Y_%m_%d", time.localtime()).split('_')
        today_date=ymd[0]+'_'+str(int(ymd[1]))+'_'+str(int(ymd[2]))
        #print today_date
        if table_date==today_date:
            self.table_name=self.mysql_table_name+'_'+table_set
        else:
            self.table_name=self.mysql_table_name+'_'+table_date+'_'+table_set
        
    def set_backup_name(self):
        backup_Set=self.config.getstr(self.mysql_table_name,'backup_Set','restore_table')
        #print backup_Set
        self.table_name=self.mysql_table_name+'_'+backup_Set
        
    def get_id_list(self):
        '''random get id list'''
        self.Key=self.config.getstr(self.mysql_table_name,'Key',"restore_table")
        self.redis_table_name=self.config.getstr(self.mysql_table_name,'redis_table_name',"restore_table")
        self.id_list= self.config.get_list(self.mysql_table_name,self.Key,"restore_table")
        
        if 'random' in self.id_list:
            id_num=self.id_list[1]
            all_id_list=[]
            random_id_list=[]
            if  self.redis_table_name=='user_info':
                delimiter='_'
                id_info_list= self.r.keys(self.redis_table_name+':uid_*')
            else:
                delimiter=':'
                id_info_list= self.r.keys(self.redis_table_name+':*')
            for id_info in id_info_list:
                all_id=id_info.split(delimiter)[-1]
                all_id_list.append(all_id)
            max_num=len(id_info_list)
            #print 'max_num',max_num
            while len(random_id_list)<int(id_num):
                random_id=random.choice(all_id_list)
                if random_id not in random_id_list:
                    random_id_list.append(random_id)
                if len(random_id_list)==max_num:
                    break
            self.id_list=random_id_list

    def restore_table(self,oa=None):
        restore_name=self.mysql_table_name
        if oa=='all':restore_name='all'
        cmd='cd /home/qihancloud/tools/data_recover/bin/;./data_recover '+restore_name+' '+self.restore_info
        #print cmd
        stdin,stdout,stderr=self.ssh.exec_command(cmd)
        result_list=stdout.readlines()
        for result in result_list:
            print result

        
    def get_count(self):
        suffix=';'
        #print self.redis_table_name
        if self.redis_table_name=='user_info':
            count_redis=len(self.r.keys(self.redis_table_name+':qlink_id_*'))+len(self.r.keys(self.redis_table_name+':device_*'))
        else:
            count_redis=len(self.r.keys(self.redis_table_name+':*'))
        if self.mysql_table_name=='file_info' and self.flag=='restore':suffix=' where flag=1 and type=1'
        elif self.mysql_table_name=='friend_info':suffix=' where flag=1 group by uid'    
        count_mysql=self.cur.execute('select * from '+self.table_name+suffix)
        #print count_mysql,count_redis
        return (count_mysql,count_redis)
    
    def sort_data(self):
        
        redis_index_list=[]
        temp_keys=[]
        temp_values=[]
        useless=['create_time','update_time','flag']
        #get redis_index in mysql
        flag=0
        for i in range(len(self.id_list)):
            redis_index_list.append([])
            for j in range(len(self.mysql_keys_list[i])):
                try:
                    index=self.redis_keys_list[i].index(self.mysql_keys_list[i][j])
                    redis_index_list[i].append(index)
                except ValueError as e:
                    if self.mysql_keys_list[i][j] not in useless:
                        if flag==0:
                            print self.mysql_keys_list[i][j],'is not in redis_keys'
                            flag=1
                        else:
                            pass
                    redis_index_list[i].append('no_key')
            a=set(self.redis_keys_list[i])
            b=set(self.mysql_keys_list[i])
            temp=list(a-b)# 求差集（项在a中，但不在b中）  
            redis_index_list[i]=redis_index_list[i]+temp    
                            
        #sort redis_list order by mysql_list  
        for i in range(len(self.id_list)):
            temp_keys.append([])
            temp_values.append([])
            for index in redis_index_list[i]:
                if index=='no_key':
                    temp_keys[i].append('no_key')
                    temp_values[i].append('no_data')
                elif type(index)==int:
                    temp_keys[i].append(self.redis_keys_list[i][index])
                    temp_values[i].append(self.redis_values_list[i][index])
                else:
                    #print self.redis_keys_list[i].index(index)
                    temp_keys[i].append(index)
                    temp_values[i].append(self.redis_values_list[i][self.redis_keys_list[i].index(index)])

        self.redis_keys_list[:]=temp_keys
        self.redis_values_list[:]=temp_values
        
    def get_data1(self):
        #select * , hgetall
        num=0;
        self.cur.execute('select * from %s limit 1'%self.table_name)
        for desc in self.cur.description:
            num=num+1
            self.mysql_keys.append(desc[0])
        for my_id in self.id_list:
            print self.table_name+":"+my_id
            self.cur.execute('select * from %s where %s=%s'%(self.table_name,self.Key,my_id))
            mysql_values=self.cur.fetchone()
            if not mysql_values:
                print self.Key+" "+my_id +" not in mysql table "+self.table_name
                mysql_values=[my_id]+["ZW1wdHk="]*(num-1)
            self.mysql_keys_list.append(self.mysql_keys)
            mysql_values=list(mysql_values)
            redis_keys=self.r.hkeys('%s:%s'%(self.redis_table_name,my_id))
            redis_values=self.r.hvals('%s:%s'%(self.redis_table_name,my_id))
            if self.mysql_table_name=='user_info':
                redis_keys.insert(0,self.Key)
                redis_values.insert(0,my_id)
                mysql_values[4]=base64.b64decode(mysql_values[4])
            elif self.mysql_table_name=='file_info':
                id_index=redis_keys.index('id')
                redis_keys[id_index]='file_id'
                mysql_values[1]=base64.b64decode(mysql_values[1])
                mysql_values[13]=base64.b64decode(mysql_values[13])
            elif self.mysql_table_name=='local_user_status':
                redis_keys.insert(0,self.Key)
                redis_values.insert(0,my_id)
                mysql_values[8]=base64.b64decode(mysql_values[8])
                mysql_values[9]=base64.b64decode(mysql_values[9])
                mysql_values[10]=base64.b64decode(mysql_values[10])
                
            self.redis_keys_list.append(redis_keys)
            self.redis_values_list.append(redis_values)
            self.mysql_values_list.append(mysql_values)   
        self.sort_data()
           
    def get_data2(self):
        
        self.cur.execute('select * from %s limit 1'%self.table_name)
        for desc in self.cur.description:self.mysql_keys.append(desc[0])
        for my_id in self.id_list:
            values_list=[]
            remarks_network_list=[]
            friend_id_list=[]
            result=self.cur.execute('select * from %s where %s=%s and flag=1'%(self.table_name,self.Key,my_id))
            for i in range(result):
                mysql_values=list(self.cur.fetchone())
                friend_id=str(mysql_values[1])
                friend_id_list.append(friend_id)
                mysql_values[2]=base64.b64decode(mysql_values[2])
                #print mysql_values[2]
                values_list.append(mysql_values)
 
            friend_uid_list= list(self.r.smembers('%s:%s'%(self.redis_table_name,my_id))) #获取好友列表
            a=set(friend_id_list)
            b=set(friend_uid_list)
            difference=list(b-a)
            for i in range(len(difference)):
                values_list.append([' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' '])
            friend_id_list=friend_id_list+difference
     
            for friend_uid in friend_id_list:
                #print friend_uid
                remarks= self.r.hmget('friend_list_remarks:'+my_id,'remarks:'+friend_uid,'remarks_version:'+friend_uid,'base_version:'+friend_uid)
                network= self.r.hmget('network_perm:'+my_id,'permission:'+friend_uid,'u64externd:'+friend_uid,'strexternd:'+friend_uid)
                remarks.insert(0,friend_uid)#插入friend_id
                remarks.insert(0,my_id) #插入uid
                if self.flag=='backup':
                    if remarks[3]==None:remarks[3]='0' # redis中remarks_version字段为空时（值为0）不显示，备份到mysql中为0
                    if remarks[4]==None:remarks[4]='0' # redis中base_version字段为空时（值为0）不显示，备份到mysql中为0
                remarks_network=remarks+network+[' ',' ',' '] #按照mysql字段，组合remarks，network
                remarks_network_list.append(remarks_network)
            self.mysql_values_list.append(values_list)
            self.redis_values_list.append(remarks_network_list)
            
    @classmethod
    def create_xlsx(cls,xlsx_path,mysql_tables_list):
       
        w=Workbook()
        ws=w.active
        ws.title='user_info'
        ws.cell(row=1,column=1).value='self.mysql_doc'
        ws.cell(row=2,column=1).value='self.mysql_keys'    
        ws.cell(row=4,column=1).value='user_info:总数：'    
        for t in mysql_tables_list:
            if t!='user_info':
                ws1=w.create_sheet(title=t)
                ws1.cell(row=1,column=1).value='self.mysql_doc'
                ws1.cell(row=2,column=1).value='self.mysql_keys'    
                ws1.cell(row=4,column=1).value=t+u'总数：'
        w.save(xlsx_path)

      
    def write_xlsx(self,xlsx_path):
	
        mysql_doc=self.config.get_list(self.mysql_table_name,self.mysql_table_name+'_doc','mysql_doc')
        count=self.get_count()
        if self.mysql_table_name=='friend_info':
            self.get_data2()
            write='write2()'
        else:
            self.get_data1()
            write='write1()'
        
        w=load_workbook(xlsx_path)
        ws1=w.get_sheet_by_name(self.mysql_table_name)
        font=Font(color='FF0000')
        alignment=Alignment(horizontal='left')
        #write title
        c_title=['mysql_keys','redis_keys','mysql_values','redis_values','']
        c_title= c_title * len(self.id_list)
        for i in range(len(c_title)):
            ws1.cell(row=i+6,column=1).value=c_title[i]
        ws1.cell(row=4,column=2).value='mysql:'+str(count[0])
        ws1.cell(row=4,column=3).value='redis:'+str(count[1])
        if count[0]!=count[1]:
            ws1.cell(row=4,column=3).font=font

        #write mysql_doc and self.mysql_keys
        while len(mysql_doc)<len(self.mysql_keys):
            mysql_doc.append(' ')    
        for i in range(len(mysql_doc)):
            try:
                #print mysql_doc[i],self.mysql_keys[i]
                ws1.cell(row=1,column=i+2).value=mysql_doc[i]
                ws1.cell(row=2,column=i+2).value=self.mysql_keys[i]
                if mysql_doc[i]!=self.mysql_keys[i]:
                    ws1.cell(row=2,column=i+2).font=font
            except IndexError as e:
                print 'IndexError,self.mysql_keys is not match mysql_doc'
                ws1.cell(row=1,column=i+2).font=font
                
        def write1():        
            #write normal data to Excel        
            values=(self.mysql_keys_list,self.redis_keys_list,self.mysql_values_list,self.redis_values_list)    
            for i in range(4):
                for j in range(len(self.id_list)):
                    for k in range(len(values[i][j])):
                        try:
                            ws1.cell(row=j*5+i+6,column=k+2).value=values[i][j][k]
                            ws1.cell(row=j*5+i+6,column=k+2).alignment=alignment
                            if i>0 and (i+1)%2==0:
                                try:
                                    if str(values[i][j][k])!=str(values[i-1][j][k]):
                                        ws1.cell(row=j*5+i+6,column=k+2).font=font
                                except IndexError as e:
                                    ws1.cell(row=j*5+i+6,column=k+2).font=font
                        except UnicodeDecodeError as e:
                            print 'UnicodeDecodeError: ',values[i][j][k]
                            ws1.cell(row=j*5+i+6,column=k+2).value='UnicodeDecodeError'
                            ws1.cell(row=j*5+i+6,column=k+2).font=font

        def write2():
            #write friend_info to excel
            start_row=7                        
            for i in range(len(self.redis_values_list)):
                flag=0
                ws1.cell(row=start_row-1,column=1).value='mysql_keys'
                for j in range(len(self.redis_values_list[i])):
                    ws1.cell(row=start_row,column=1).value='mysql_values'
                    ws1.cell(row=start_row+1,column=1).value='redis_values'
                    for k in range(len(self.redis_values_list[i][j])):
                        try:
                            if flag==0:
                                ws1.cell(row=start_row-1,column=k+2).value=self.mysql_keys[k]
                            ws1.cell(row=start_row,column=k+2).value=str(self.mysql_values_list[i][j][k])
                            ws1.cell(row=start_row+1,column=k+2).value=self.redis_values_list[i][j][k]
                            ws1.cell(row=start_row+1,column=k+2).alignment=alignment
                            if str(self.mysql_values_list[i][j][k])!=self.redis_values_list[i][j][k]:
                                ws1.cell(row=start_row+1,column=k+2).font=font
                        except UnicodeDecodeError as e:
                            #print e
                            print 'UnicodeDecodeError',self.redis_values_list[i][j][k]
                    flag=1
                    start_row=start_row+2
                start_row=start_row+2

        eval(write)
        w.save(xlsx_path)

       
        
